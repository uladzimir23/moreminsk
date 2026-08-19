#!/usr/bin/env python3
"""Dry-run parser: xlsx календарь → CSV для ревью перед импортом в PB."""
import re
import csv
import openpyxl

SRC = "/Users/vladimirmazyrec/Downloads/ЛЕТО 2026.xlsx"
OUT = "/tmp/bookings-dryrun.csv"

MONTHS = {
    "Май": 5, "Июнь": 6, "Июль": 7,
    "Август": 8, "Сентябрь": 9, "Октябрь": 10,
}
YACHTS = {"EVA", "ALFA", "BRAVO", "MARIO", "КУРАЖ"}

# Регексы для свободного текста
# Строгие форматы: BY = +375 + 9 цифр, RU = +7/8 + 10 цифр. Не жадные.
RE_PHONE_BY = re.compile(r"\+?375[\s\-\(\)]*\d{2}[\s\-\(\)]*\d{3}[\s\-]?\d{2}[\s\-]?\d{2}")
RE_PHONE_RU = re.compile(r"(?:\+?7|8)[\s\-\(\)]*\d{3}[\s\-\(\)]*\d{3}[\s\-]?\d{2}[\s\-]?\d{2}")
RE_GUESTS = re.compile(r"(\d{1,2})\s*(?:чел|человек|xtk|взр|детей)", re.IGNORECASE)
# Цена: standalone 2-4-значное число (50..900), опционально с (NN) prepaid.
# Ищем в тексте БЕЗ телефона (телефон вырезаем заранее).
RE_PRICE = re.compile(r"(?<!\d)(\d{2,4})\s*(?:\(\s*(\d{1,4})\s*\))?(?!\d)")
RE_TIME_RANGE = re.compile(r"(\d{1,2})[.:](\d{2})\s*[-–]\s*(\d{1,2})[.:](\d{2})")
RE_DAY = re.compile(r"^(\d{1,2})\s*\(", re.UNICODE)
RE_START_ONLY = re.compile(r"(\d{1,2})[.:](\d{2})\s+[А-ЯЁA-Z]")  # "10.30 Константин" — half-hour start
# Хвосты которые режем из name
NAME_TAIL_RE = re.compile(r"\s*(?:\d+\s*(?:чел|xtk|взр|детей|человек|ч|р|рб|руб)|\d+[-–]\d+).*$", re.IGNORECASE)

AGENCIES = {"юность", "юности", "сливки", "рцоп", "робинсон"}
PAY_HINTS = {"безнал", "бн", "аванс", "нал", "перекатать"}


def month_from_sheet(name: str) -> int | None:
    for k, v in MONTHS.items():
        if name.strip().lower().startswith(k.lower()):
            return v
    return None


def norm_yacht(v: str) -> str | None:
    if v is None:
        return None
    s = str(v).strip().upper()
    if s in YACHTS:
        return s.capitalize() if s in ("MARIO", "КУРАЖ") else s
    return None


def normalize_phone(raw: str) -> str:
    """+375 (29) 194-00-35 → +375291940035 (E.164-ish)."""
    digits = re.sub(r"[^\d+]", "", raw)
    if digits.startswith("8") and len(digits) == 11:
        digits = "+7" + digits[1:]
    elif digits.startswith("7") and len(digits) == 11:
        digits = "+" + digits
    elif digits.startswith("375"):
        digits = "+" + digits
    return digits


def parse_cell(text: str) -> dict:
    """Извлечь name / phone / guests / price / pay_note / source из свободного текста."""
    if not text or not text.strip():
        return {}
    raw = text.strip()
    cleaned = raw  # Будем последовательно вырезать распознанные куски

    # Phone: сначала BY, потом RU. Нормализуем в E.164.
    phone = ""
    m = RE_PHONE_BY.search(cleaned) or RE_PHONE_RU.search(cleaned)
    if m:
        phone = normalize_phone(m.group(0))
        cleaned = cleaned.replace(m.group(0), " ")

    # Guests
    guests = 0
    g_m = RE_GUESTS.search(cleaned)
    if g_m:
        guests = int(g_m.group(1))
        cleaned = cleaned.replace(g_m.group(0), " ")

    # Price + prepaid — теперь в тексте без телефона
    price = 0
    prepaid = 0
    for pm in RE_PRICE.finditer(cleaned):
        val = int(pm.group(1))
        if 50 <= val <= 900:
            price = val
            if pm.group(2):
                pval = int(pm.group(2))
                if 30 <= pval <= 500:
                    prepaid = pval
            break

    # Разделяем: агентство-партнёр (source) vs платёжный хинт (pay_note)
    low = raw.lower()
    source_agents = sorted({a for a in AGENCIES if a in low})
    pay_hints = sorted({h for h in PAY_HINTS if h in low})

    # Name: первая непустая строка с буквой, чистим хвост «N чел / N р / N-N»
    name = ""
    for line in raw.split("\n"):
        line = line.strip()
        if not line or line.isdigit():
            continue
        if RE_PHONE_BY.match(line) or RE_PHONE_RU.match(line):
            continue
        if not re.search(r"[А-Яа-яA-Za-z]", line):
            continue
        # обрезаем до первой запятой / точки / телефона
        n = re.split(r"[,\.]", line, maxsplit=1)[0].strip()
        n = RE_PHONE_BY.sub("", n)
        n = RE_PHONE_RU.sub("", n)
        n = NAME_TAIL_RE.sub("", n).strip()
        # выкидываем trailing symbols/дефисы
        n = re.sub(r"[-\s]+$", "", n).strip()
        if n:
            name = n
            break

    return {
        "name": name,
        "phone": phone,
        "guests": guests,
        "price": price,
        "prepaid": prepaid,
        "pay_note": ",".join(pay_hints),
        "source_agent": ",".join(source_agents),
    }


def parse_sheet(ws, month: int, year: int = 2026) -> list[dict]:
    """Найти week-блоки: строка с яхтами → collect map col→yacht → 12 рядов ниже = слоты."""
    rows = list(ws.iter_rows(values_only=False))
    out = []

    for i, row in enumerate(rows):
        # Ищем «yacht header row» — строка где ≥3 ячейки = EVA/ALFA/BRAVO/...
        yacht_cols = {}
        for cell in row:
            y = norm_yacht(cell.value)
            if y:
                yacht_cols[cell.column] = y
        if len(yacht_cols) < 3:
            continue

        # Есть yacht-row на индексе i. Ищем day-header row выше (обычно i-1)
        # В day-row: cells где текст `N (день)` → маппинг col → date
        day_cols = {}  # start_col → day_number
        for lookback in (1, 2):
            if i - lookback < 0:
                break
            for cell in rows[i - lookback]:
                if cell.value is None:
                    continue
                m = RE_DAY.match(str(cell.value).strip())
                if m:
                    day_cols[cell.column] = int(m.group(1))
            if day_cols:
                break

        if not day_cols:
            continue

        # Для каждого дня определим range колонок: [start_col, next_start_col)
        day_ranges = []
        sorted_days = sorted(day_cols.items())
        for k, (col, day) in enumerate(sorted_days):
            end_col = sorted_days[k + 1][0] if k + 1 < len(sorted_days) else max(yacht_cols.keys()) + 2
            day_ranges.append((col, end_col, day))

        # Идём вниз до 12 hour-rows (или пока не встретим следующий week-блок)
        slot_row_count = 0
        j = i + 1
        while j < len(rows) and slot_row_count < 14:
            r = rows[j]
            # Найти time-cell в этой строке
            time_range = None
            half_hour_start = None
            for cell in r:
                if cell.value is None:
                    continue
                s = str(cell.value).strip()
                m = RE_TIME_RANGE.search(s)
                if m and cell.column <= max(yacht_cols.keys()):
                    time_range = (
                        f"{int(m.group(1)):02d}:{m.group(2)}",
                        f"{int(m.group(3)):02d}:{m.group(4)}",
                    )
                    break
            if not time_range:
                # Пустая строка между блоками = конец
                if all(c.value is None for c in r):
                    if slot_row_count > 0:
                        break
                j += 1
                continue

            slot_row_count += 1
            start_time, end_time = time_range

            # Идём по каждой ячейке в этой строке; определяем yacht/day по колонке
            for cell in r:
                if cell.value is None:
                    continue
                val = str(cell.value).strip()
                if not val or RE_TIME_RANGE.search(val):
                    continue

                # Определяем yacht: ближайший yacht_cols в этой же day-range
                # yacht_cols keys — колонки где стоит имя яхты
                yacht = None
                for yc, yn in yacht_cols.items():
                    if yc == cell.column:
                        yacht = yn
                        break
                if not yacht:
                    continue

                # Определяем date по day_ranges
                day = None
                for (col_start, col_end, d) in day_ranges:
                    if col_start <= cell.column < col_end:
                        day = d
                        break
                if not day:
                    continue

                date_str = f"{year:04d}-{month:02d}-{day:02d}"

                # Half-hour start hint внутри текста
                real_start = start_time
                half_m = RE_START_ONLY.search(val)
                if half_m:
                    hh = int(half_m.group(1))
                    mm = half_m.group(2)
                    if f"{hh:02d}" == start_time[:2] and mm == "30":
                        real_start = f"{hh:02d}:30"

                parsed = parse_cell(val)
                out.append({
                    "month": f"{year}-{month:02d}",
                    "date": date_str,
                    "yacht": yacht,
                    "start": real_start,
                    "end": end_time,
                    "name": parsed.get("name", ""),
                    "phone": parsed.get("phone", ""),
                    "guests": parsed.get("guests", 0),
                    "price": parsed.get("price", 0),
                    "prepaid": parsed.get("prepaid", 0),
                    "pay_note": parsed.get("pay_note", ""),
                    "source_agent": parsed.get("source_agent", ""),
                    "raw": re.sub(r"\s+", " ", val).strip()[:200],
                })
            j += 1

    return out


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    all_rows = []
    for name in wb.sheetnames:
        month = month_from_sheet(name)
        if not month:
            continue
        ws = wb[name]
        rows = parse_sheet(ws, month)
        print(f"[{name}] found {len(rows)} filled slots")
        all_rows.extend(rows)

    # Дедуп по (date, yacht, start, raw)
    seen = set()
    dedup = []
    for r in all_rows:
        k = (r["date"], r["yacht"], r["start"], r["raw"])
        if k in seen:
            continue
        seen.add(k)
        dedup.append(r)

    # Сортируем: date, yacht, start
    dedup.sort(key=lambda r: (r["date"], r["yacht"], r["start"]))

    with open(OUT, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=[
            "date", "yacht", "start", "end",
            "name", "phone", "guests", "price", "prepaid",
            "source_agent", "pay_note", "raw",
        ])
        w.writeheader()
        for r in dedup:
            w.writerow({k: r[k] for k in [
                "date", "yacht", "start", "end",
                "name", "phone", "guests", "price", "prepaid",
                "source_agent", "pay_note", "raw",
            ]})

    print(f"\nTOTAL rows (deduped): {len(dedup)} → {OUT}")

    # Stats
    by_month = {}
    with_phone = 0
    with_price = 0
    with_name = 0
    empty_all = 0
    for r in dedup:
        by_month[r["month"]] = by_month.get(r["month"], 0) + 1
        if r["phone"]:
            with_phone += 1
        if r["price"]:
            with_price += 1
        if r["name"]:
            with_name += 1
        if not (r["phone"] or r["price"] or r["name"]):
            empty_all += 1

    print(f"\nBy month:")
    for m, c in sorted(by_month.items()):
        print(f"  {m}: {c}")
    print(f"\nCoverage:")
    print(f"  with phone: {with_phone} ({with_phone*100//len(dedup)}%)")
    print(f"  with price: {with_price} ({with_price*100//len(dedup)}%)")
    print(f"  with name:  {with_name} ({with_name*100//len(dedup)}%)")
    print(f"  raw-only (no name/phone/price parsed): {empty_all}")


if __name__ == "__main__":
    main()
